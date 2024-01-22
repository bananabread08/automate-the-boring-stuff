import openpyxl

workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))

# get a sheet
sheet = workbook.get_sheet_by_name('Sheet1')
print(type(sheet))

print(workbook.sheetnames)

# get a cell by location
value = sheet['A1'].value
print(value)


# get a cell with row and column provided
cell = sheet.cell(row=1, column=2).value # Apples
print(cell)

# show all column values by looping
for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)


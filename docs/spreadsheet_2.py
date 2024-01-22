import openpyxl

# open sheet in-memory
wb = openpyxl.Workbook()

sheet = wb['Sheet']
sheet['A1'] = 42

# create another sheet
new_sheet = wb.create_sheet()
new_sheet.title = 'Another Sheet'


# create sheet on a specific index
wb.create_sheet(index=0, title='My Other Sheet')

print(wb.sheetnames)

# save this as test.xlsx
wb.save('test.xlsx')
